#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import sys
import glob
import shutil
import openpyxl
import json
import threading
from datetime import datetime
import logging
from logging.handlers import TimedRotatingFileHandler

import tkinter as tk
import tkinter.messagebox as messagebox
from tkinter import filedialog


class WinGUI(object):
    """docstring for App"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("提取网关发货清单")
        # self.root.iconbitmap('csg.ico')
        self.root.geometry("800x400")
        self.root.minsize(800, 400)
        self.root.bind("<Configure>", self.on_resize)
        self.__dst_path = tk.StringVar()
        self.__src_path = tk.StringVar()
        self.create_main_window()
        self.config = self.load_config()
        self.__src_path.set(self.config.get('src_path', ''))
        self.__dst_path.set(self.config.get('dst_path', ''))
        self.thread_list = []

    @property
    def src_path(self):
        return self.__src_path.get()

    @property
    def dst_path(self):
        return self.__dst_path.get()

    def create_main_window(self):

        select_label1 = tk.Label(self.root, text='发货文件夹')
        select_label1.grid(row=0, column=0, sticky='w', padx=10, pady=5)
        select_label2 = tk.Label(self.root, text='输出文件夹')
        select_label2.grid(row=1, column=0, sticky='w', padx=10, pady=5)

        select_entry1 = tk.Entry(self.root, textvariable=self.__src_path, width=40)
        select_entry1.grid(row=0, column=1, sticky='e', padx=10, pady=5)

        select_button1 = tk.Button(self.root, text="选择发货目录", command=self.__on_select_src, width=10)
        select_button1.grid(row=0, column=2, sticky='w')

        select_entry2 = tk.Entry(self.root, textvariable=self.__dst_path, width=40)
        select_entry2.grid(row=1, column=1, sticky='e', padx=10, pady=5)

        select_button2 = tk.Button(self.root, text="选择输出目录", command=self.__on_select_dst, width=10)
        select_button2.grid(row=1, column=2, sticky='w')

        run_button = tk.Button(self.root, text="运行", command=self.on_run, width=10)
        run_button.grid(row=0, column=3, rowspan=3, sticky='nsew', padx=10, pady=5)

        # 滚动条
        scrollbar = tk.Scrollbar(self.root)
        scrollbar.grid(row=3, column=4, sticky='nsew')

        # 日志文本框
        self.text_widget = tk.Text(self.root, state="disabled", wrap=tk.WORD, yscrollcommand=scrollbar.set)
        self.text_widget.grid(row=3, column=0, columnspan=4, sticky='nsew')
        scrollbar.config(command=self.text_widget.yview)

        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(3, weight=1)

    def __on_select_src(self):
        select_path = filedialog.askdirectory()
        if select_path:
            self.__src_path.set(select_path)
        self.save_config('src_path', select_path)
        self.print_log(f'源发货文件夹路径为：{select_path}')

    def __on_select_dst(self):
        select_path = filedialog.askdirectory()
        if select_path:
            self.__dst_path.set(select_path)
        self.save_config('dst_path', select_path)
        self.print_log(f'输出文件夹路径为：{select_path}')

    def run(self):
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()

    def on_run(self):
        if not os.path.exists(self.src_path):
            self.print_log('发货清单源文件夹路径不存在')
            return False
        if not os.path.exists(self.dst_path):
            self.print_log('需要生成合并发货清单的文件夹路径不存在')
            return False
        if self.is_run:
            self.print_log('已经在运行了')
            return 0
        self.save_config('src_path', self.src_path)
        self.save_config('dst_path', self.dst_path)
        t = threading.Thread(target=merge_excel)
        t.daemon = True
        t.start()
        self.is_run = t

    @property
    def is_run(self):
        ret = False
        for item in self.thread_list:
            if item.is_alive():
                ret = True
                break
        return ret

    @is_run.setter
    def is_run(self, t):
        self.thread_list.append(t)

    def on_closing(self):
        if not self.is_run:
            self.root.destroy()
            exit()
        if messagebox.askokcancel("退出", "当前正在处理excel合并任务，退出可能会导致excel文件出现错误，确定退出吗?"):
            self.root.destroy()
            exit()

    def on_resize(self, event):
        try:
            self.text_widget.config(width=event.width)
        except:
            pass

    def print_log(self, text):
        now = datetime.now()
        date = now.strftime("%Y-%m-%d %H:%M:%S")
        self.text_widget.config(state="normal")
        logger.info(text)
        self.text_widget.insert("end", f"{date}：{text}\n")
        self.text_widget.yview_moveto(1)
        self.text_widget.see("end")
        if self.text_widget.index('end-1c') == '1000.0':
            self.text_widget.delete(1.0, 2.0)
        self.text_widget.config(state="disabled")

    @staticmethod
    def load_config():
        config = {}
        if os.path.isfile("config.json"):
            with open("config.json", "r") as f:
                config = json.load(f)
        return config

    def save_config(self, key, value):
        config = self.load_config()
        config[key] = value
        with open("config.json", "w") as f:
            json.dump(config, f)


def int_to_excel_col(col: int) -> str:
    # 将数字列数转成excel的字母列数
    col_str = ""
    while col > 0:
        col -= 1
        col, rem = divmod(col, 26)
        col_str = chr(rem + ord('A')) + col_str
    return col_str


def excel_col_to_int(col_letter):
    # 将 Excel 列字母转换为列数字
    col_number = 0
    for i, letter in enumerate(col_letter):
        # 从 A 开始，每个字母的权值为 26 的幂次方
        col_number += (ord(letter.upper()) - 64) * (26 ** (len(col_letter) - i - 1))
    return col_number


def get_all_file(folder_path):
    # 搜索文件夹下的所有Excel和Word文件
    excel_files = [f for f in glob.glob(os.path.join(folder_path, '*.xls*')) if
                   not os.path.basename(f).startswith('~$')]
    return excel_files


def get_file_type(file_path):
    file_ext = os.path.splitext(file_path)[1].lower()

    if file_ext == '.xls' or file_ext == '.xlsx':
        return 'Excel'
    elif file_ext == '.doc' or file_ext == '.docx':
        return 'Word'
    else:
        return 'Unknown'


def unmerge_cells(sheet, cell):
    '''
    拆分单元格，将单元格的数据从sheet中清除，并将单元格的值设置为第一个值
    
    :param sheet: 工作表
    :param cell: 单元格
    '''
    r1, r2, c1, c2 = cell.min_row, cell.max_row, cell.min_col, cell.max_col
    # 拆分合并单元格
    sheet.unmerge_cells(start_row=r1, end_row=r2, start_column=c1, end_column=c2)
    # 在所有拆分出的单元格内填充数据
    first_value = sheet.cell(r1, c1).value
    for r in range(r1, r2 + 1):  # 遍历行
        if c2 - c1 > 0:  # 多个列，遍历列
            for c in range(c1, c2 + 1):
                sheet.cell(r, c).value = first_value
        else:  # 一个列
            sheet.cell(r, c1).value = first_value


def get_is_merged(sheet, cell):
    '''
    判断单元格是否被合并
    :param sheet: 所在的sheet
    :param cell: 单元格
    :return: 如果被合并返回合并单元格的范围，否则返回None
    '''
    ret = None
    merged_cells_ranges = sheet.merged_cells.ranges
    for merged_range in merged_cells_ranges:
        if cell.coordinate in merged_range:
            ret = merged_range
            break
    return ret


def get_shipping_info(sheet, dic):
    print('*' * 30)
    ret = {}
    # 遍历关键词
    for key in dic.keys():
        find_str = dic.get(key)
        # 遍历表格所有的行
        for row in sheet.iter_rows():
            if ret.get(find_str):
                # 如果关键词已经找到，就该关键词的循环
                break
            # 遍历改行所有的单元格
            for cell in row:
                # 如果单元格包含关键词
                if find_str in str(cell.value):
                    # 判断这个单元格是否是合并的单元格
                    key_cell_range = get_is_merged(sheet, cell)
                    # 如果是合并的单元格
                    if key_cell_range:
                        # 获取合并单元格的位置信息
                        r1, r2, c1, c2 = key_cell_range.min_row, key_cell_range.max_row, key_cell_range.min_col, key_cell_range.max_col
                        # 目标单元格为合并单元的的最小行和最大列+1
                        traget_cell = sheet.cell(row=r1, column=c2 + 1)
                    else:
                        # 如果不是合并单元格，目标单元格为该单元格的列+1，由于行的单元格列表索引从0开始，所以不用+1
                        traget_cell = row[cell.column]
                    # 判断目标单元格是否是合并的单元格，如果是则拆分单元格
                    merged_range = get_is_merged(sheet, traget_cell)
                    if merged_range:
                        unmerge_cells(sheet, merged_range)
                    # 获取目标单元格的值，并存到字典中
                    ret[key] = traget_cell.value
                    print(f'{key}:找到了，值是:{traget_cell.value}')
                    # 跳过该行的循环
                    break
    return ret


def get_table_range(sheet, start, end):
    # 按关键词获取内容的范围
    start_cell_coordinate = None
    end_cell_coordinate = None
    # 遍历sheet中的每一行
    for row in sheet.iter_rows():
        # 如果start_cell_coordinate和end_cell_coordinate都不为空，则跳出循环
        if start_cell_coordinate and end_cell_coordinate:
            break
        # 遍历每一行中的每一列
        for cell in row:
            # 获取每一列的值
            cell_value = str(cell.value)
            # 如果start在每一列的值中，则将start_cell_coordinate赋值给start_cell_coordinate
            if start in cell_value:
                start_cell_coordinate = cell.coordinate
            # 如果end在每一列的值中，则将end_cell_coordinate赋值给end_cell_coordinate
            if end in cell_value:
                end_cell_coordinate = cell.coordinate

    # 返回start_cell_coordinate和end_cell_coordinate
    return start_cell_coordinate, end_cell_coordinate


def get_header(rows, shipping_list_key):
    dic = {}
    for row in rows:
        for key in shipping_list_key:
            for cell in row:
                if key in str(cell.value):
                    dic[key] = cell.column
    return dic


def get_shipping_list(sheet, shipping_list_key):
    ret = []
    # 获取发货清单的范围
    start_str, end_str = get_table_range(sheet, start='序号', end='合计')
    if not end_str:
        return None
    start_cell = sheet[start_str]
    end_cell = sheet[end_str]
    header_row = sheet.iter_rows(min_row=start_cell.row, max_row=start_cell.row)
    header_dic = get_header(header_row, shipping_list_key)
    # 按关键词获取表头的列存入字典
    for row in sheet.iter_rows(min_row=start_cell.row + 1, max_row=end_cell.row - 1):
        row_data = {}
        for key in shipping_list_key:
            for cell in row:
                merged_range = get_is_merged(sheet, cell)
                if merged_range:
                    unmerge_cells(sheet, merged_range)
            try:
                row_data[key] = sheet.cell(row=cell.row, column=header_dic[key]).value
            except:
                row_data[key] = ''
                gui.print_log(f'该文件中没有{key}列')
        ret.append(row_data)

    return ret


def merge_excel():
    gui.print_log('开始'.center(30, '='))
    gui.print_log(f'原始发货清单路径：{gui.src_path}')
    gui.print_log(f'合并后的发货清单路径：{gui.dst_path}')
    date = datetime.now().strftime("%Y-%m-%d_%H_%M_%S")
    out_save_path = os.path.join(gui.dst_path, f'发货清单_{date}.xlsx')
    err_dst_dir = os.path.join(gui.src_path, f'程序处理不了的发货清单文件_{date}')
    excel_files = get_all_file(gui.src_path)
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    row = out_ws.max_row
    all_gw_count = 0
    header_col = main_config.get('table_header')
    if not header_col:
        gui.print_log(f'加载配置文件\"{config_file}\"时出现问题，请检查该文件是否存在或格式是否正确!')
        gui.print_log(f'因为缺少必要的配置，程序停止执行!')
        return False
    for src_file in excel_files:
        print(excel_files)
        gui.print_log(f'处理文件：{src_file}')
        try:
            wb = openpyxl.load_workbook(src_file)
            sheet = wb.active
        except:
            wb.close()
            gui.print_log(f'在处理文件\"{src_file}\"时出现未知错误！')
            continue
        shipping_info = get_shipping_info(sheet, shipping_info_dic)
        shipping_list = get_shipping_list(sheet, shipping_list_key)
        if not shipping_list:
            gui.print_log(f'该文件格式超出预期，无法找到发货清单区域,预期的发货清单数据在"序号"--"合计"之间：{src_file}')
            wb.close()
            if not os.path.exists(err_dst_dir):
                os.makedirs(err_dst_dir)
            shutil.copy(src_file, err_dst_dir)
            continue
        gw_count = 0
        for item in shipping_list:
            item.update(shipping_info)
            for key in header_col.keys():
                out_ws[f'{header_col[key]}{row}'] = item.get(key)
            row += 1
            gw_count += 1
        gui.print_log(f'该文件有{gw_count}条数据')
        wb.close()
        all_gw_count += 1
        if all_gw_count % 10 == 0:
            out_wb.save(out_save_path)
    out_wb.save(out_save_path)
    out_wb.close()
    gui.print_log(f'文件已保存到{out_save_path}')
    gui.print_log(f'本次共处理发货清单文件{all_gw_count}个')
    gui.print_log(f'全部任务已完成')


def setup_logging(log_file='app.log'):
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # 创建一个文件handler
    file_handler = TimedRotatingFileHandler(log_file, when="midnight", interval=1, backupCount=7, encoding='utf-8')
    file_handler.setFormatter(
        logging.Formatter('%(asctime)s - %(name)s - %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(file_handler)

    # 创建一个控制台handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.DEBUG)
    console_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - [%(filename)s:%('
                                                   'lineno)d] %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(console_handler)

if __name__ == '__main__':
    if getattr(sys, 'frozen', False):
        PATH = os.path.dirname(sys.executable)
    elif __file__:
        PATH = os.path.split(os.path.realpath(__file__))[0]
    sys.path.append(os.path.join(PATH, 'libs', 'site-packages'))
    # 日志
    now = datetime.now()
    date = now.strftime("%Y-%m-%d_%H_%M_%S")
    log_dir = os.path.join(PATH, 'logs')
    if not os.path.exists(log_dir):
        os.mkdir(log_dir)
    log_path = os.path.join(PATH, 'logs', f'日志_{date}.log')
    setup_logging(log_path)
    logger = logging.getLogger()
    gui = WinGUI()
    config_file = 'mainconf.json'
    main_config = {}
    try:
        with open(os.path.join(PATH, config_file), 'r', encoding='utf-8') as f:
            json_str = f.read()
            main_config = json.loads(json_str)

    except:
        gui.print_log(f'加载配置文件"{config_file}"时出现问题，请检查该文件是否存在或格式是否正确!')

    try:
        # 获取当前用户主目录路径
        user_home = os.path.expanduser('~')

        # 拼接获取用户目录路径
        user_path = os.path.join(user_home, 'Documents')
        with open(os.path.join(user_path, config_file), 'r', encoding='utf-8') as f:
            json_str = f.read()
            main_config = json.loads(json_str)
        gui.print_log(f'已在{user_path}找到配置文件"{config_file}"!')
    except:
        pass

    shipping_info_dic = main_config.get('shipping_info')
    shipping_list_key = main_config.get('shipping_list_key')

    if not shipping_info_dic:
        shipping_info_dic = {
            "合同编号": "合同编号",
            "省": "建设单位（省）",
            "项目负责人": "项目负责人",
            "市": "市",
            "县": "县",
            "收货人": "收货人",
            "收货单位": "收货单位",
            "收货人电话": "收货人电话",
            "收货地址": "收货地址",
            "发货通知人": "发货通知人",
            "通知人电话": "通知人电话",
            "出库人": "出库人",
            "出库人电话": "出库人电话",
            "发货人": "发货人",
            "出库日期": "出库日期",
            "发货单位": "发货单位",
            "发货地址": "发货地址",
            "批次": "批次"
        }
    if not shipping_list_key:
        shipping_list_key = (
            '序号', '物资编码', '产品名称', '规格型号', 'SN编号', '单位', '数量', '项目名称', '项目编号', '备注',
            'ICCID')


    gui.run()
